#Este arquivo Python utiliza a seguinte codificação: utf-8

#Importando bibliotecas
import os
import time
import datetime
from datetime import datetime
import openpyxl
import requests
from tkinter import *
from openpyxl import Workbook, load_workbook

#Criando banco de dados
try:
    funcionarios = load_workbook('Funcionários.xlsx')
    dados = funcionarios.active
except:
    funcionarios = Workbook()
    dados = funcionarios.active
    dados.title = "Funcionários"
    dados.append(['Código', 'Nome', 'Cargo', 'CPF', 'Salário', 'Horário de entrada', 'Horário de saída', 'Última data registrada', 'Última hora registrada'])
    funcionarios.save('Funcionários.xlsx')

#Carregando banco de dados
funcionarios = load_workbook('Funcionários.xlsx')
dados = funcionarios.active
#print(dados['A1'].value) #Imprime valor da célula A1 contendo "Código"

#Declarando funções
def situacaoFuncionario(info):
    global situacao_atual
    if(info == "Entrada"):
        situacao_atual['text'] = "Situação: Ativo"
    elif(info == "Pausa"):
        situacao_atual['text'] = "Situação: Em pausa"
    elif(info == "Retorno pausa"):
        situacao_atual['text'] = "Situação: Ativo"
    elif(info == "Saída"):
        situacao_atual['text'] = "Situação: Inativo"

#Receber data e horário do sistema
def horarioDoSistema():
    global horario
    global hora_atual
    hora_atual = time.ctime().split()[3]
    horario.config(text="Hora atual: {}".format(hora_atual))
    horario.after(1000, horarioDoSistema)

def ultimaHoraRegistrada():
    global en_codigo
    global hora_registrada
    global data_registrada
    funcionarios = load_workbook('Funcionários.xlsx')
    dados = funcionarios.active
    for row in dados.iter_rows():
        for cell in row:
            if str(cell.value) != str(en_codigo.get()):
                continue
            else:    
                #print("Data atual é {}.".format(str(dados.cell(row=cell.row, column=1).value)))
                hora = dados.cell(row=cell.row, column=9).value
                hora_registrada['text'] = "Última hora registrada: {}".format(hora)
                data = dados.cell(row=cell.row, column=8).value
                data_registrada['text'] = "Última data registrada: {}".format(data)

def registrarHora():
    global en_codigo
    funcionarios = load_workbook('Funcionários.xlsx')
    dados = funcionarios.active
    hora = hora_atual
    for row in dados.iter_rows():
        for cell in row:
            if str(cell.value) != str(en_codigo.get()):
                continue
            else:    
                #print("Data atual é {}.".format(str(dados.cell(row=cell.row, column=1).value)))
                dados.cell(row=cell.row, column=9).value = hora
    funcionarios.save("Funcionários.xlsx")

def registrarData():
    global en_codigo
    funcionarios = load_workbook('Funcionários.xlsx')
    dados = funcionarios.active
    hora = hora_atual
    for row in dados.iter_rows():
        for cell in row:
            if str(cell.value) != str(en_codigo.get()):
                continue
            else:    
                #print("Data atual é {}.".format(str(dados.cell(row=cell.row, column=1).value)))
                dados.cell(row=cell.row, column=8).value = datetime.today().strftime('%Y-%m-%d')
    funcionarios.save("Funcionários.xlsx")

def Entrada(): #Função entrada
    global entrada
    global pausa
    global retorno_pausa
    global saida
    global cancelar_ponto
    entrada['state'] = DISABLED
    pausa['state'] = ACTIVE
    saida['state'] = ACTIVE
    cancelar_ponto['state'] = DISABLED
    info = "Entrada"
    hora = hora_atual
    relatorio(info, hora)
    situacaoFuncionario(info)

def Pausa(): #Função pausa
    global pausa
    global retorno_pausa
    pausa['state'] = DISABLED
    retorno_pausa['state'] = ACTIVE
    info = "Pausa"
    hora = hora_atual
    relatorio(info, hora)
    situacaoFuncionario(info)

def RetornoPausa(): #Função retorno pausa
    global retorno_pausa
    global pausa
    retorno_pausa['state'] = DISABLED
    pausa['state'] = ACTIVE
    info = "Retorno pausa"
    hora = hora_atual
    relatorio(info, hora)
    situacaoFuncionario(info)

def Saida(): #Função saída
    global saida
    global entrada
    global janela
    saida['state'] = DISABLED
    entrada['state'] = ACTIVE
    info = "Saída"
    hora = hora_atual
    relatorio(info, hora)
    situacaoFuncionario(info)
    janela.destroy()
    janelaPrincipal()

#Função para verificar se o mês mudou
def mesAtual():
    global funcionarios
    global en_codigo
    dados = funcionarios.active
    data_atual = datetime.today().strftime('%Y-%m-%d')
    data_atual = data_atual.split("-")
    mes_atual = int(data_atual[1])
    #print(data_atual)
    #print(datetime.today().strftime('%Y-%m-%d'))
    #print(mes_atual)
    for row in dados.iter_rows():
            for cell in row:
                if str(cell.value) != str(en_codigo.get()):
                    continue
                else:    
                    data_registrada = str(dados.cell(row=cell.row, column=8).value)
                    mes_registrado = data_registrada.split("-")[1]
                    #print(mes_registrado)
                    #print(data_registrada)
                    if(int(mes_registrado) != int(mes_atual)):
                        registrarData()
                        i = False
                    else:
                        i = True
    if(i == False):
        #print("Mês diferente do atual")
        return False
    else:
        #print("Mês retornado é o atual")
        return True

def diaAtual():
    global funcionarios
    global en_codigo
    dados = funcionarios.active
    data_atual = datetime.today().strftime('%Y-%m-%d')
    data_atual = data_atual.split("-")
    dia_atual = int(data_atual[0])
    for row in dados.iter_rows():
            for cell in row:
                if str(cell.value) != str(en_codigo.get()):
                    continue
                else:    
                    data_registrada = str(dados.cell(row=cell.row, column=8).value)
                    dia_registrado = data_registrada.split("-")[0]
                    if(int(dia_registrado) != int(dia_atual)):
                        i = False
                    else:
                        i = True
    if(i == False):
        return False
    else:
        return True

#Função que gera os relatórios mensais dos funcionários
def relatorio(info, hora):
    global nome
    mesAtual()
    diaAtual()
    data_atual = datetime.today().strftime('%Y-%m-%d')
    data_atual = data_atual.split("-")
    mes_atual = int(data_atual[1])
    global planilha
    planilha = "{} - Mês {}".format(nome , mes_atual)
    #print(mesAtual())
    if (mesAtual() == False): #Criando nova planilha mensal
        relatorio = Workbook()
        dados = relatorio.active
        dados.title = "{}".format(planilha)
        dados.append(['Data', 'Entrada', 'Pausa', 'Retorno pausa', 'Saída', 'Atraso', 'Extra'])
        dados.append([datetime.today().strftime('%Y-%m-%d')])
        for row in dados.iter_rows():
            for cell in row:
                if str(cell.value) != str(datetime.today().strftime('%Y-%m-%d')):
                    continue
                else:    
                    #print("Data atual é {}.".format(str(dados.cell(row=cell.row, column=1).value)))
                    if(info == "Entrada"):
                        dados.cell(row=cell.row, column=2).value = hora #Inserindo horário de entrada atual no relatório
                        registrarHora()
                        registrarData()
                        ultimaHoraRegistrada()
                    elif(info == "Pausa"):
                        dados.cell(row=cell.row, column=3).value = hora #Inserindo horário de pausa atual no relatório
                        registrarHora()
                        registrarData()
                        ultimaHoraRegistrada()
                    elif(info == "Retorno pausa"):
                        dados.cell(row=cell.row, column=4).value = hora #Inserindo horário de retorno da pausa atual no relatório
                        registrarHora()
                        registrarData()
                        ultimaHoraRegistrada()
                    elif(info == "Saída"):
                        dados.cell(row=cell.row, column=5).value = hora #Inserindo horário de saída atual no relatório
                        registrarHora()
                        registrarData()
                        ultimaHoraRegistrada()
        relatorio.save('{}.xlsx'.format(planilha))
        if (info == "Entrada"):
            dados.cell(row=cell.row, column=6).value = horasAtrasadas()
            relatorio.save('{}.xlsx'.format(planilha))
        elif (info == "Saída"):
            dados.cell(row=cell.row, column=7).value = horasExtras()
            relatorio.save('{}.xlsx'.format(planilha))

        #Gerando relatório mensal do funcionário
        

    elif(mesAtual() == True): #Atualizando planilha do mês atual
        try:
            relatorio = load_workbook("{}.xlsx".format(planilha))
            dados = relatorio.active
            if(diaAtual() == False):
                dados.append([datetime.today().strftime('%Y-%m-%d')])
            for row in dados.iter_rows():
                for cell in row:
                    if str(cell.value) != str(datetime.today().strftime('%Y-%m-%d')):
                        continue
                    else:    
                        #print("Data atual é {}.".format(str(dados.cell(row=cell.row, column=1).value)))
                        if(info == "Entrada"):
                            dados.cell(row=cell.row, column=2).value = hora #Inserindo horário de entrada atual no relatório
                            registrarHora()
                            registrarData()
                            ultimaHoraRegistrada()
                        elif(info == "Pausa"):
                            dados.cell(row=cell.row, column=3).value = hora #Inserindo horário de pausa atual no relatório
                            registrarHora()
                            registrarData()
                            ultimaHoraRegistrada() 
                        elif(info == "Retorno pausa"):
                            dados.cell(row=cell.row, column=4).value = hora #Inserindo horário de retorno da pausa atual no relatório
                            registrarHora()
                            registrarData()
                            ultimaHoraRegistrada()
                        elif(info == "Saída"):
                            dados.cell(row=cell.row, column=5).value = hora #Inserindo horário de saída atual no relatório
                            registrarHora()
                            registrarData()
                            ultimaHoraRegistrada()
            relatorio.save('{}.xlsx'.format(planilha))
            if (info == "Entrada"):
                dados.cell(row=cell.row, column=6).value = horasAtrasadas()
                relatorio.save('{}.xlsx'.format(planilha))
            elif (info == "Saída"):
                dados.cell(row=cell.row, column=7).value = horasExtras()
                relatorio.save('{}.xlsx'.format(planilha))
        except:
            relatorio = Workbook()
            dados = relatorio.active
            dados.title = "{}".format(planilha)
            dados.append(['Data', 'Entrada', 'Pausa', 'Retorno pausa', 'Saída', 'Atraso', 'Extra'])
            dados.append([datetime.today().strftime('%Y-%m-%d')])
            for row in dados.iter_rows():
                for cell in row:
                    if str(cell.value) != str(datetime.today().strftime('%Y-%m-%d')):
                        continue
                    else:    
                        #print("Data atual é {}.".format(str(dados.cell(row=cell.row, column=1).value)))
                        if(info == "Entrada"):
                            dados.cell(row=cell.row, column=2).value = hora #Inserindo horário de entrada atual no relatório
                            registrarHora()
                            registrarData()
                            ultimaHoraRegistrada()
                        elif(info == "Pausa"):
                            dados.cell(row=cell.row, column=3).value = hora #Inserindo horário de pausa atual no relatório
                            registrarHora()
                            registrarData()
                            ultimaHoraRegistrada()
                        elif(info == "Retorno pausa"):
                            dados.cell(row=cell.row, column=4).value = hora #Inserindo horário de retorno da pausa atual no relatório
                            registrarHora()
                            registrarData()
                            ultimaHoraRegistrada()
                        elif(info == "Saída"):
                            dados.cell(row=cell.row, column=5).value = hora #Inserindo horário de saída atual no relatório
                            registrarHora()
                            registrarData()
                            ultimaHoraRegistrada()
            relatorio.save('{}.xlsx'.format(planilha))
            if (info == "Entrada"):
                dados.cell(row=cell.row, column=6).value = horasAtrasadas()
                relatorio.save('{}.xlsx'.format(planilha))
            elif (info == "Saída"):
                dados.cell(row=cell.row, column=7).value = horasExtras()
                relatorio.save('{}.xlsx'.format(planilha))

def lerHoraData():
    relatorio = load_workbook("{}.xlsx".format(planilha))
    dados = relatorio.active
    for row in dados.iter_rows():
        for cell in row:
            if str(cell.value) != str(datetime.today().strftime('%Y-%m-%d')):
                continue
            else:    
                #print("Data atual é {}.".format(str(dados.cell(row=cell.row, column=1).value)))
                entrada = dados.cell(row=cell.row, column=2).value
                pausa = dados.cell(row=cell.row, column=3).value
                retorno_pausa = dados.cell(row=cell.row, column=4).value
                saida = dados.cell(row=cell.row, column=5).value
    return entrada, pausa, retorno_pausa, saida
    
#Função do ponto eletrônico propriamente dito
def pontoEletronico():
    global janela
    global cadastrar
    global horario
    horario.destroy()
    cadastrar.destroy()
    global situacao_atual
    situacao_atual = Label(janela, text="Situação: Inativo", font = ("Century Gothic", 12))
    situacao_atual.grid(column = 0, row = 4, padx = 10, pady = 10)
    global hora_registrada
    hora_registrada = Label(janela, text="", font = ("Century Gothic", 12))
    hora_registrada.grid(column = 0, row = 5, padx = 10, pady = 10)
    global data_registrada
    data_registrada = Label(janela, text="", font = ("Century Gothic", 12))
    data_registrada.grid(column = 0, row = 6, padx = 10, pady = 10)
    global entrada
    entrada = Button(janela, text="Entrada", font = ("Century Gothic", 12), command = lambda: [Entrada()]) #Botão para registrar entrada
    entrada.grid(column = 1, row = 4, padx = 10, pady = 10)
    global pausa
    pausa = Button(janela, text="Pausa", font = ("Century Gothic", 12), state = DISABLED, command = lambda: [Pausa()]) #Botão para registrar pausa
    pausa.grid(column = 2, row = 4, padx = 10, pady = 10)
    global retorno_pausa
    retorno_pausa = Button(janela, text="Retorno pausa", font = ("Century Gothic", 12), state = DISABLED, command = lambda: [RetornoPausa()]) #Botão para registrar retorno da pausa
    retorno_pausa.grid(column = 1, row = 5, padx = 10, pady = 10)
    global saida
    saida = Button(janela, text="Saída", font = ("Century Gothic", 12), state = DISABLED, command = lambda: [Saida()]) #Botão para registrar saída
    saida.grid(column = 2, row = 5, padx = 10, pady = 10)
    horario = Label(janela, text="", font = ("Century Gothic", 12))
    horario.grid(column = 1, row = 7, padx = 10, pady = 10)
    global cancelar_ponto
    cancelar_ponto = Button(janela, text="Cancelar", font = ("Century Gothic", 12), command = lambda: [janela.destroy(), janelaPrincipal()])
    cancelar_ponto.grid(column = 0, row = 7, padx = 10, pady = 10)
    horarioDoSistema()
    ultimaHoraRegistrada()

def janelaCadastro():
    #Criando janela de cadastro
    formulario = Tk()
    formulario.title("Ponto Eletrônico")

    #Inserindo textos na janela
    orientacao = Label(formulario, text="Preencha todos os campos.", font = ("Century Gothic", 12))
    orientacao.grid(column = 0, row = 0, padx = 10, pady = 10)

    txt_nome = Label(formulario, text="Digite o nome", font = ("Century Gothic", 12))
    txt_nome.grid(column = 0, row = 1, padx = 10, pady = 10)
    
    txt_cargo = Label(formulario, text="Digite o cargo", font = ("Century Gothic", 12))
    txt_cargo.grid(column = 0, row = 3, padx = 10, pady = 10)

    txt_cpf = Label(formulario, text="Digite o CPF", font = ("Century Gothic", 12))
    txt_cpf.grid(column = 0, row = 5, padx = 10, pady = 10)
    
    txt_salario = Label(formulario, text="Digite o salário mensal (R$)", font = ("Century Gothic", 12))
    txt_salario.grid(column = 1, row = 1, padx = 10, pady = 10)    
    
    txt_hora_entrada = Label(formulario, text="Digite a hora de entrada (hh:mm:ss)", font = ("Century Gothic", 12))
    txt_hora_entrada.grid(column = 1, row = 3, padx = 10, pady = 10)
    
    txt_hora_saida = Label(formulario, text="Digite a hora de saída (hh:mm:ss)", font = ("Century Gothic", 12))
    txt_hora_saida.grid(column = 1, row = 5, padx = 10, pady = 10)
    
    #Criando entrada de dados na janela
    en_nome = Entry(formulario, border = 2, font = ("Century Gothic", 12), justify = CENTER)
    en_nome.grid(column = 0, row = 2, padx = 10, pady = 10)

    en_cargo = Entry(formulario, border = 2, font = ("Century Gothic", 12), justify = CENTER)
    en_cargo.grid(column = 0, row = 4, padx = 10, pady = 10)

    en_cpf = Entry(formulario, border = 2, font = ("Century Gothic", 12), justify = CENTER)
    en_cpf.grid(column = 0, row = 6, padx = 10, pady = 10)

    en_salario = Entry(formulario, border = 2, font = ("Century Gothic", 12), justify = CENTER)
    en_salario.grid(column = 1, row = 2, padx = 10, pady = 10)

    en_hora_entrada = Entry(formulario, border = 2, font = ("Century Gothic", 12), justify = CENTER)
    en_hora_entrada.grid(column = 1, row = 4, padx = 10, pady = 10)

    en_hora_saida = Entry(formulario, border = 2, font = ("Century Gothic", 12), justify = CENTER)
    en_hora_saida.grid(column = 1, row = 6, padx = 10, pady = 10)

    #Criando botões na janela
    cadastrar = Button(formulario, text="Finalizar cadastro", font = ("Century Gothic", 12), command = lambda: [cadastrarFuncionario(en_nome, en_cargo, en_salario, en_cpf, en_hora_entrada, en_hora_saida), formulario.destroy()])
    cadastrar.grid(column = 0, row = 7, padx = 10, pady = 10)

    cancelar_cadastro = Button(formulario, text="Cancelar", font = ("Century Gothic", 12), command = lambda: [formulario.destroy()])
    cancelar_cadastro.grid(column = 1, row = 7, padx = 10, pady = 10)

    #Mantendo janela aberta
    formulario.mainloop()

def cadastrarFuncionario(en_nome, en_cargo, en_salario, en_cpf, en_hora_entrada, en_hora_saida):
    global funcionarios
    dados = funcionarios.active
    global resultado
    codigo = dados.max_row
    #print(en_nome.get())
    dados.append([codigo, en_nome.get(), en_cargo.get(), en_cpf.get(), en_salario.get(), en_hora_entrada.get(), en_hora_saida.get(), datetime.today().strftime('%Y-%m-%d'), time.ctime().split()[3]])
    funcionarios.save('Funcionários.xlsx')
    resultado['text'] = "O seu código de identificação é {}. Assegure-se de guardá-lo ou memorizá-lo.".format(codigo)
    data_atual = datetime.today().strftime('%Y-%m-%d')
    data_atual = data_atual.split("-")
    mes_atual = int(data_atual[1])
    planilha = "{} - Mês {}".format(en_nome.get() , mes_atual)
    relatorio = Workbook()
    dados = relatorio.active
    dados.title = "{}".format(planilha)
    dados.append(['Data', 'Entrada', 'Pausa', 'Retorno pausa', 'Saída'])
    dados.append([datetime.today().strftime('%Y-%m-%d')])
    relatorio.save('{}.xlsx'.format(planilha))

def localizarFuncionario(en_codigo):
    global funcionarios
    dados = funcionarios.active
    i = False
    #Procurando código do funcionário no banco de dados
    for row in dados.iter_rows():
        for cell in row:
            if str(cell.value) != str(en_codigo.get()):
                continue
            else:    
                #print("Número encontrado é {}.".format(str(dados.cell(row=cell.row, column=1).value)))
                global nome
                nome = dados.cell(row=cell.row, column=2).value
                cargo = dados.cell(row=cell.row, column=3).value
                cpf = dados.cell(row=cell.row, column=4).value
                salario = dados.cell(row=cell.row, column=5).value
                hora_entrada = dados.cell(row=cell.row, column=6).value
                hora_saida = dados.cell(row=cell.row, column=7).value
                i = True
    if(i == False):
        funcionario = "Funcionário não identificado. Tente novamente ou realize o cadastro."
    else:    
        funcionario = "Funcionário identificado. Parabéns!\nNome: {}\nCargo: {}\nCPF: {}\nSalário: R$ {}\nHorário de entrada: {}\nHorário de saída: {}".format(nome, cargo, cpf, salario, hora_entrada, hora_saida)
        pontoEletronico()
        orientacao.grid_remove()
        en_codigo.grid_remove()
        identificar.grid_remove()
    return funcionario

def identificarFuncionario(en_codigo):
    global resultado
    resposta = localizarFuncionario(en_codigo)
    resultado["text"] = f'{resposta}'

#Define horário de entrada do funcionário
#def horarioEntrada()

#Define horário de saída do funcionário
#def horarioSaida()

#Calcula horas adiantadas em relação ao horário de entrada
#def horasAdiantadas()

def contarSegundos(hora, min, seg):
    segundos = hora*3600 + min*60 + seg
    return segundos

def segundosParaHoras(segundos):
    return time.strftime("%H:%M:%S", time.gmtime(segundos)) 

#Calcula horas atrasadas em relação ao horário de entrada
def horasAtrasadas():
    global funcionarios
    dados = funcionarios.active
    i = False
    #Procurando código do funcionário no banco de dados
    for row in dados.iter_rows():
        for cell in row:
            if str(cell.value) != str(en_codigo.get()):
                continue
            else:    
                #print("Número encontrado é {}.".format(str(dados.cell(row=cell.row, column=1).value)))
                hora_entrada = dados.cell(row=cell.row, column=6).value
                hora_saida = dados.cell(row=cell.row, column=7).value
    entrada, pausa, retorno_pausa, saida = lerHoraData()
    hora_entrada = hora_entrada.split(":")
    entrada = entrada.split(":")
    hora_entrada_segundos = contarSegundos(int(hora_entrada[0]), int(hora_entrada[1]), int(hora_entrada[2]))
    entrada_segundos = contarSegundos(int(entrada[0]), int(entrada[1]), int(entrada[2]))
    atraso = entrada_segundos - hora_entrada_segundos
    if(atraso < 0):
        return "00:00:00"
    else: 
        return segundosParaHoras(atraso)

#Calcula horas trabalhadas
#def horasTrabalhadas()

#Calcula horas extras em relação ao horário de saída
def horasExtras():
    global funcionarios
    dados = funcionarios.active
    i = False
    #Procurando código do funcionário no banco de dados
    for row in dados.iter_rows():
        for cell in row:
            if str(cell.value) != str(en_codigo.get()):
                continue
            else:    
                #print("Número encontrado é {}.".format(str(dados.cell(row=cell.row, column=1).value)))
                hora_entrada = dados.cell(row=cell.row, column=6).value
                hora_saida = dados.cell(row=cell.row, column=7).value
    entrada, pausa, retorno_pausa, saida = lerHoraData()
    hora_saida = hora_saida.split(":")
    saida = saida.split(":")
    hora_saida_segundos = contarSegundos(int(hora_saida[0]), int(hora_saida[1]), int(hora_saida[2]))
    saida_segundos = contarSegundos(int(saida[0]), int(saida[1]), int(saida[2]))
    extra = saida_segundos - hora_saida_segundos
    if(extra < 0):
        return "00:00:00"
    else:
        return segundosParaHoras(extra)

def janelaPrincipal():
    global janela
    #Criando janela principal
    janela = Tk()
    janela.title("Ponto Eletrônico")
    #Declarando variáveis
    esconder_codigo = StringVar()

    #Importando imagens

    #Inserindo textos na janela
    global orientacao
    orientacao = Label(janela, text="Digite o código do funcionário.", font = ("Century Gothic", 12))
    orientacao.grid(column = 1, row = 0, padx = 10, pady= 10)

    global resultado
    resultado = Label(janela, text="", font = ("Century Gothic", 12))
    resultado.grid(column = 1, row = 3, padx= 10, pady = 10)

    global horario
    horario = Label(janela, text="", font = ("Century Gothic", 12))
    horario.grid(column = 1, row = 5, padx= 10, pady = 10)
    horarioDoSistema()

    #Criando entrada de dados da janela
    global en_codigo
    en_codigo = Entry(janela, textvariable=esconder_codigo, show="*", border=2, font = ("Century Gothic", 12), justify=CENTER)
    en_codigo.grid(column = 1, row = 1, padx= 10, pady = 10)

    #Inserindo botões na janela
    global cadastrar
    cadastrar = Button(janela, text="Cadastrar funcionário", font = ("Century Gothic", 12), command = lambda: [janelaCadastro()])
    cadastrar.grid(column = 1, row = 4, padx = 10, pady = 10)

    global identificar
    identificar = Button(janela, text="Identificar funcionário", font = ("Century Gothic", 12), command = lambda: identificarFuncionario(en_codigo))
    identificar.grid(column = 1, row = 2, padx = 10, pady = 10)

    #Mantendo janela aberta
    janela.mainloop()

janelaPrincipal()